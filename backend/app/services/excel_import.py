from __future__ import annotations

import csv
import io
from decimal import Decimal, InvalidOperation

from openpyxl import Workbook, load_workbook
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.catalog import Category, TruckBrand, TruckModel
from app.models.enums import ProductStatus
from app.models.product import Product, ProductVehicle

# Column order for the import template / expected headers
TEMPLATE_COLUMNS = [
    "category",  # slug or Russian name
    "truck_brand",  # name or slug (optional)
    "model",  # optional
    "name_ru",
    "name_uz",
    "article",
    "oem_number",
    "part_brand",
    "price",
    "stock",
    "warranty",
    "engine",
    "description_ru",
    "description_uz",
    "image_urls",  # comma-separated URLs (optional)
]


class ImportError_(ValueError):
    pass


def build_template_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "products"
    ws.append(TEMPLATE_COLUMNS)
    ws.append(
        [
            "filtry",
            "MAN",
            "TGX",
            "Фильтр масляный",
            "Moy filtri",
            "51.10100-6126",
            "51.10100-6126",
            "MANN",
            "150000",
            "20",
            "6 мес",
            "D2066",
            "Оригинальный масляный фильтр",
            "Original moy filtri",
            "",
        ]
    )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _rows_from_bytes(data: bytes, filename: str) -> list[dict]:
    name = (filename or "").lower()
    if name.endswith(".csv"):
        text = data.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        return [dict(r) for r in reader]
    # default: xlsx
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    out: list[dict] = []
    for raw in rows[1:]:
        if raw is None or all(c is None for c in raw):
            continue
        out.append({headers[i]: raw[i] if i < len(raw) else None for i in range(len(headers))})
    return out


def _to_decimal(value) -> Decimal:
    if value is None or str(value).strip() == "":
        return Decimal("0")
    try:
        return Decimal(str(value).replace(" ", "").replace(",", "."))
    except InvalidOperation as exc:
        raise ImportError_("invalid_price") from exc


def _to_int(value) -> int:
    if value is None or str(value).strip() == "":
        return 0
    try:
        return int(float(str(value).replace(" ", "")))
    except ValueError as exc:
        raise ImportError_("invalid_stock") from exc


async def import_products(
    session: AsyncSession,
    seller_id: int,
    data: bytes,
    filename: str,
    auto_approve: bool = False,
) -> dict:
    rows = _rows_from_bytes(data, filename)
    if not rows:
        return {"created": 0, "errors": [{"row": 0, "error": "empty_file"}]}

    # Preload categories and brands for fast lookup
    categories = (await session.execute(select(Category))).scalars().all()
    cat_by_slug = {c.slug.lower(): c for c in categories}
    cat_by_name = {c.name_ru.lower(): c for c in categories}
    cat_by_name_uz = {c.name_uz.lower(): c for c in categories}

    brands = (await session.execute(select(TruckBrand))).scalars().all()
    brand_by_key = {}
    for b in brands:
        brand_by_key[b.name.lower()] = b
        brand_by_key[b.slug.lower()] = b

    created = 0
    errors: list[dict] = []
    status = ProductStatus.approved if auto_approve else ProductStatus.pending

    for idx, row in enumerate(rows, start=2):  # header is row 1
        try:
            cat_key = str(row.get("category") or "").strip().lower()
            category = (
                cat_by_slug.get(cat_key) or cat_by_name.get(cat_key) or cat_by_name_uz.get(cat_key)
            )
            if category is None:
                raise ImportError_(f"unknown_category:{cat_key}")

            name_ru = str(row.get("name_ru") or "").strip()
            if not name_ru:
                raise ImportError_("missing_name_ru")

            product = Product(
                seller_id=seller_id,
                category_id=category.id,
                name_ru=name_ru,
                name_uz=str(row.get("name_uz") or "").strip(),
                article=str(row.get("article") or "").strip(),
                oem_number=str(row.get("oem_number") or "").strip(),
                part_brand=str(row.get("part_brand") or "").strip(),
                engine=str(row.get("engine") or "").strip(),
                description_ru=str(row.get("description_ru") or "").strip(),
                description_uz=str(row.get("description_uz") or "").strip(),
                price=_to_decimal(row.get("price")),
                stock_qty=_to_int(row.get("stock")),
                warranty=str(row.get("warranty") or "").strip(),
                status=status,
                is_active=True,
            )
            session.add(product)
            await session.flush()

            # Optional compatible vehicle
            brand_key = str(row.get("truck_brand") or "").strip().lower()
            if brand_key and brand_key in brand_by_key:
                brand = brand_by_key[brand_key]
                model_id = None
                model_name = str(row.get("model") or "").strip()
                if model_name:
                    existing_model = (
                        await session.execute(
                            select(TruckModel).where(
                                TruckModel.brand_id == brand.id,
                                func.lower(TruckModel.name) == model_name.lower(),
                            )
                        )
                    ).scalar_one_or_none()
                    if existing_model is None:
                        existing_model = TruckModel(brand_id=brand.id, name=model_name)
                        session.add(existing_model)
                        await session.flush()
                    model_id = existing_model.id
                session.add(
                    ProductVehicle(
                        product_id=product.id, truck_brand_id=brand.id, truck_model_id=model_id
                    )
                )
            created += 1
        except ImportError_ as exc:
            errors.append({"row": idx, "error": str(exc)})
        except Exception as exc:  # noqa: BLE001 - report row-level failures, keep going
            errors.append({"row": idx, "error": f"unexpected:{exc}"})

    await session.flush()
    return {"created": created, "errors": errors}
